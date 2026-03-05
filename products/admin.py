from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils import timezone
from django.urls import reverse, path
from django.shortcuts import redirect, get_object_or_404
from django.db.models import Count, Sum, Q, F
from django.template.response import TemplateResponse
from django.db.models import PROTECT as PROTECT_ON_DELETE
from mptt.admin import DraggableMPTTAdmin
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime, date
from exports.mixins import ConfigurableExportMixin
from exports.column_config import ColumnConfigMixin
from rbac.admin_mixins import ERPAdminMixin
from .models import (
    Unit, ProductCategory, Product, ProductPacking, 
    ProductImage, ProductAttribute, ProductAttributeValue,
    ProductAttributeAssignment, ProductVariant, ProductPriceHistory,
    Brand, ModelNumber, PriceList, ProductPrice
)
from . import views




# -----------------------------
# Excel Export Mixin
# -----------------------------
class ExcelExportMixin:
    """Mixin to add Excel export functionality to admin"""
    
    def export_as_excel(self, request, queryset):
        """Export selected items as Excel file"""
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{self.model.__name__} Export"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Get fields to export
            fields = self.get_export_fields()
            
            # Write headers
            for col, field in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = str(field).upper()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, obj in enumerate(queryset, 2):
                for col, field in enumerate(fields, 1):
                    try:
                        value = self.get_field_value(obj, field)
                        cell = ws.cell(row=row, column=col, value=value)
                        
                        # Format based on data type
                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            cell.number_format = '#,##0.00'
                        elif isinstance(value, datetime):
                            cell.number_format = 'YYYY-MM-DD HH:MM'
                    except Exception as e:
                        print(f"Error exporting {field} for {obj}: {e}")
                        ws.cell(row=row, column=col, value="ERROR")
            
            # Auto-size columns
            for col in range(1, len(fields) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, min(queryset.count() + 2, 100)):
                    try:
                        cell_value = ws[f"{column_letter}{row}"].value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{self.model.__name__}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            wb.save(response)
            return response
            
        except Exception as e:
            self.message_user(request, f"❌ Export failed: {str(e)}", level='ERROR')
            return redirect(request.path)
    
    export_as_excel.short_description = "Export selected as Excel"
    
    def get_export_fields(self):
        """Define which fields to export - override in each admin"""
        if hasattr(self, 'export_fields'):
            return self.export_fields
        return [field.name for field in self.model._meta.fields]
    
    def get_field_value(self, obj, field_name):
        """Get field value handling relationships, methods, and timezone issues"""
        try:
            # Check if it's a method in export_methods
            if hasattr(self, 'export_methods') and field_name in self.export_methods:
                return self.export_methods[field_name](obj)
            
            # Handle related fields with __
            if '__' in field_name:
                parts = field_name.split('__')
                value = obj
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        break
                return self._prepare_value_for_excel(value)
            
            # Handle regular fields
            value = getattr(obj, field_name, '')
            return self._prepare_value_for_excel(value)
            
        except Exception as e:
            return f"Error: {str(e)}"
    
    def _prepare_value_for_excel(self, value):
        """Prepare any value for Excel export, handling timezone issues"""
        if value is None:
            return ""
        
        # Handle datetime objects - REMOVE TIMEZONE INFO
        if isinstance(value, datetime):
            if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            return value
        
        # Handle date objects
        if isinstance(value, date):
            return value
        
        # Handle decimal/float
        if isinstance(value, Decimal):
            return float(value)
        
        # Handle boolean
        if isinstance(value, bool):
            return "Yes" if value else "No"
        
        # Handle related objects
        if hasattr(value, '__str__'):
            return str(value)
        
        return value


# -----------------------------
# Custom Filters
# -----------------------------
class HasVariantsFilter(admin.SimpleListFilter):
    title = 'Has Variants'
    parameter_name = 'has_variants'

    def lookups(self, request, model_admin):
        return (
            ('yes', '✅ Yes - Has Variants'),
            ('no', '❌ No - No Variants'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(variants__isnull=False).distinct()
        if self.value() == 'no':
            return queryset.filter(variants__isnull=True)
        return queryset


# REMOVED: StockStatusFilter (no longer needed)


class PriceRangeFilter(admin.SimpleListFilter):
    title = 'Price Range'
    parameter_name = 'price_range'

    def lookups(self, request, model_admin):
        return (
            ('0-25', '💰 Under $25'),
            ('25-50', '💰 $25 to $50'),
            ('50-100', '💰 $50 to $100'),
            ('100-500', '💰 $100 to $500'),
            ('500-1000', '💰 $500 to $1000'),
            ('1000+', '💰 Over $1000'),
        )

    def queryset(self, request, queryset):
        if self.value() == '0-25':
            return queryset.filter(price__lt=25)
        if self.value() == '25-50':
            return queryset.filter(price__gte=25, price__lt=50)
        if self.value() == '50-100':
            return queryset.filter(price__gte=50, price__lt=100)
        if self.value() == '100-500':
            return queryset.filter(price__gte=100, price__lt=500)
        if self.value() == '500-1000':
            return queryset.filter(price__gte=500, price__lt=1000)
        if self.value() == '1000+':
            return queryset.filter(price__gte=1000)
        return queryset


# -----------------------------
# Price List Admin
# -----------------------------
@admin.register(PriceList)
class PriceListAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    ALL_LIST_COLUMNS = [
        ('name', 'Name'),
        ('code', 'Code'),
        ('priority', 'Priority'),
        ('discount_method', 'Discount Method'),
        ('product_count', 'Products'),
        ('is_active', 'Active'),
        ('is_default', 'Default'),
    ]
    DEFAULT_COLUMNS = ['name', 'code', 'priority', 'discount_method', 'product_count', 'is_active', 'is_default']
    REQUIRED_COLUMNS = ['name']
    list_display = ('name', 'code', 'priority', 'discount_method', 'product_count', 'is_active', 'is_default')
    list_filter = ('is_active', 'is_default', 'discount_method')
    search_fields = ('name', 'code', 'description')
    list_editable = ('priority', 'is_active', 'is_default')
    prepopulated_fields = {'code': ('name',)}
    readonly_fields = ('created_at', 'updated_at', 'product_count')  # product_count is a method, that's fine
    list_per_page = 25
    export_fields = [
        'name', 'code', 'priority', 'discount_method', 'is_active', 'is_default', 'created_at',
    ]
    export_methods = {
        'discount_method': lambda obj: obj.get_discount_method_display(),
        'is_active': lambda obj: 'Yes' if obj.is_active else 'No',
        'is_default': lambda obj: 'Yes' if obj.is_default else 'No',
    }

    fieldsets = (
        (None, {
            'fields': ('name', 'code', 'description', 'priority')
        }),
        ('Pricing Method', {
            'fields': ('discount_method', 'default_discount_percentage')
        }),
        ('Applicable To', {
            'fields': ('applicable_to_retail', 'applicable_to_wholesale', 'applicable_to_distributor')
        }),
        ('Status', {
            'fields': ('is_active', 'is_default')
        }),
        ('Statistics', {
            'fields': ('product_count',),
            'classes': ('collapse',)
        }),
        ('Audit', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def product_count(self, obj):
        """Count products using this price list"""
        return obj.product_prices.count()
    product_count.short_description = 'Products'
    


# -----------------------------
# Product Price Inline
# -----------------------------
class ProductPriceInline(admin.TabularInline):
    model = ProductPrice
    extra = 1
    fields = ('price_list', 'price', 'min_quantity', 'valid_from', 'valid_to', 'is_valid_display')
    readonly_fields = ('is_valid_display',)
    autocomplete_fields = ['price_list']
    
    def is_valid_display(self, obj):
        from django.utils.safestring import mark_safe
        if obj.pk and obj.is_valid():
            return mark_safe('<span style="color: green;">✅ Valid</span>')
        elif obj.pk:
            return mark_safe('<span style="color: orange;">⏸️ Expired</span>')
        return mark_safe('<span style="color: gray;">⚪ Not set</span>')
    is_valid_display.short_description = 'Status'


# -----------------------------
# Unit Admin
# -----------------------------
@admin.register(Unit)
class UnitAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    list_display = ('name', 'short_name', 'code', 'unit_type', 'is_active', 'used_in_products')
    list_filter = ('unit_type', 'is_active')
    search_fields = ('name', 'short_name', 'code', 'unit_type')
    list_editable = ('is_active',)
    list_display_links = ('name', 'short_name')
    list_per_page = 25
    ALL_LIST_COLUMNS = [
        ('name', 'Name'),
        ('short_name', 'Short Name'),
        ('code', 'Code'),
        ('unit_type', 'Unit Type'),
        ('is_active', 'Active'),
        ('used_in_products', 'Used in Products'),
        ('description_preview', 'Description'),
    ]
    DEFAULT_COLUMNS = ['name', 'short_name', 'code', 'unit_type', 'is_active', 'used_in_products']
    REQUIRED_COLUMNS = ['name']
    actions = ['make_active', 'make_inactive', 'reassign_and_delete']
    save_on_top = False
    export_fields = ['name', 'short_name', 'code', 'unit_type', 'is_active', 'description']
    export_methods = {
        'unit_type': lambda obj: obj.get_unit_type_display(),
        'is_active': lambda obj: 'Yes' if obj.is_active else 'No',
    }

    fieldsets = (
        (None, {
            'fields': ('name', 'short_name', 'code', 'unit_type', 'is_active', 'description')
        }),
    )

    # ------------------------------------------------------------------
    # URLs – adds the custom reassign-and-delete intermediate page
    # ------------------------------------------------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'reassign-unit/',
                self.admin_site.admin_view(self.reassign_unit_view),
                name='products_unit_reassign',
            ),
        ]
        return custom_urls + urls

    # ------------------------------------------------------------------
    # Queryset – annotate with product counts to avoid N+1
    # ------------------------------------------------------------------
    def get_queryset(self, request):
        return super().get_queryset(request).annotate(
            _base_count=Count('base_products', distinct=True),
            _packing_count=Count('packing_units', distinct=True),
        )

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        # Autocomplete dropdowns must only offer active units
        if '/autocomplete/' in request.path:
            queryset = queryset.filter(is_active=True)
        return queryset, use_distinct

    # ------------------------------------------------------------------
    # List-display helper
    # ------------------------------------------------------------------
    def used_in_products(self, obj):
        total = obj._base_count + obj._packing_count
        if total == 0:
            return 0
        url = reverse('admin:products_product_changelist') + f'?base_unit__id__exact={obj.pk}'
        return format_html('<a href="{}">{}</a>', url, total)
    used_in_products.short_description = "Used in Products"
    used_in_products.admin_order_field = '_base_count'

    def description_preview(self, obj):
        if not obj.description:
            return '-'
        if len(obj.description) > 60:
            return obj.description[:60] + '…'
        return obj.description
    description_preview.short_description = 'Description'

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------
    def make_active(self, request, queryset):
        will_change = queryset.filter(is_active=False).count()
        already_active = queryset.count() - will_change
        queryset.update(is_active=True)
        parts = []
        if will_change:
            parts.append(f"{will_change} activated")
        if already_active:
            parts.append(f"{already_active} already active")
        self.message_user(request, ", ".join(parts) + ".")
    make_active.short_description = "Mark selected as active"

    def make_inactive(self, request, queryset):
        will_change = queryset.filter(is_active=True).count()
        already_inactive = queryset.count() - will_change
        queryset.update(is_active=False)
        parts = []
        if will_change:
            parts.append(f"{will_change} deactivated")
        if already_inactive:
            parts.append(f"{already_inactive} already inactive")
        self.message_user(request, ", ".join(parts) + ".")
    make_inactive.short_description = "Mark selected as inactive"

    def reassign_and_delete(self, request, queryset):
        """Redirect to the intermediate reassign-and-delete page."""
        ids = ','.join(str(pk) for pk in queryset.values_list('pk', flat=True))
        url = reverse('admin:products_unit_reassign') + f'?ids={ids}'
        return redirect(url)
    reassign_and_delete.short_description = "Reassign all references, then delete selected"

    # ------------------------------------------------------------------
    # Intermediate reassign-and-delete view
    # ------------------------------------------------------------------
    def reassign_unit_view(self, request):
        """
        GET : show a form listing the units to delete + a replacement dropdown.
        POST: reassign every PROTECT FK on those units to the chosen replacement,
              then delete them.
        """
        from django.db import transaction

        # Parse unit IDs from query-string (GET) or hidden field (POST).
        raw_ids = request.GET.get('ids') or request.POST.get('unit_ids', '')
        try:
            unit_ids = [int(i) for i in raw_ids.split(',') if i.strip()]
        except ValueError:
            self.message_user(request, "Invalid unit IDs.", level=messages.ERROR)
            return redirect(reverse('admin:products_unit_changelist'))

        units_to_delete = (
            Unit.objects.filter(pk__in=unit_ids)
            .annotate(
                base_count=Count('base_products', distinct=True),
                packing_count=Count('packing_units', distinct=True),
            )
        )
        if not units_to_delete.exists():
            self.message_user(request, "No matching units found.", level=messages.ERROR)
            return redirect(reverse('admin:products_unit_changelist'))

        available_units = Unit.objects.exclude(pk__in=unit_ids).order_by('unit_type', 'name')
        error = None

        if request.method == 'POST':
            replacement_id = request.POST.get('replacement_unit', '').strip()
            if not replacement_id:
                error = "Please select a replacement unit before continuing."
            else:
                try:
                    replacement = Unit.objects.get(pk=replacement_id)
                except Unit.DoesNotExist:
                    error = "The selected replacement unit no longer exists."

            if not error:
                with transaction.atomic():
                    # Dynamically find every PROTECT FK pointing at Unit and reassign.
                    for rel in Unit._meta.related_objects:
                        if rel.on_delete is PROTECT_ON_DELETE:
                            related_model = rel.related_model
                            field_name = rel.field.name
                            related_model.objects.filter(
                                **{f'{field_name}__in': unit_ids}
                            ).update(**{field_name: replacement})

                    deleted_names = list(units_to_delete.values_list('name', flat=True))
                    count = units_to_delete.count()
                    units_to_delete.delete()   # ProductPacking rows CASCADE automatically

                names_str = ', '.join(deleted_names)
                self.message_user(
                    request,
                    f"Reassigned all references to '{replacement}' and deleted "
                    f"{count} unit(s): {names_str}.",
                    level=messages.SUCCESS,
                )
                return redirect(reverse('admin:products_unit_changelist'))

        context = {
            **self.admin_site.each_context(request),
            'title': 'Reassign and delete units',
            'units_to_delete': units_to_delete,
            'unit_ids': raw_ids,
            'available_units': available_units,
            'opts': self.model._meta,
            'app_label': self.model._meta.app_label,
            'error': error,
        }
        return TemplateResponse(
            request,
            'admin/products/unit/reassign_unit.html',
            context,
        )


# -----------------------------
# Brand Admin
# -----------------------------
@admin.register(Brand)
class BrandAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    ALL_LIST_COLUMNS = [
        ('name', 'Name'),
        ('slug', 'Slug'),
        ('is_active', 'Active'),
        ('is_featured', 'Featured'),
        ('product_count', 'Products'),
        ('website_link', 'Website'),
    ]
    DEFAULT_COLUMNS = ['name', 'slug', 'is_active', 'is_featured', 'product_count', 'website_link']
    REQUIRED_COLUMNS = ['name']
    list_display = ('name', 'slug', 'is_active', 'is_featured', 'product_count', 'website_link')
    list_filter = ('is_active', 'is_featured', 'created_at')
    search_fields = ('name', 'description')
    prepopulated_fields = {'slug': ('name',)}
    readonly_fields = ('created_at', 'updated_at', 'product_count')  # product_count is a method, that's fine
    list_editable = ('is_active', 'is_featured')
    list_per_page = 25
    actions = ['make_active', 'make_inactive', 'make_featured', 'remove_featured']
    export_fields = ['name', 'slug', 'website', 'is_active', 'is_featured', 'created_at']
    export_methods = {
        'is_active': lambda obj: 'Yes' if obj.is_active else 'No',
        'is_featured': lambda obj: 'Yes' if obj.is_featured else 'No',
    }

    fieldsets = (
        (None, {
            'fields': ('name', 'slug', 'logo', 'website', 'description')
        }),
        ('Status', {
            'fields': ('is_active', 'is_featured')
        }),
        ('Statistics', {
            'fields': ('product_count',),
            'classes': ('collapse',)
        }),
        ('Audit', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def website_link(self, obj):
        if obj.website:
            return format_html('<a href="{}" target="_blank">{}</a>', obj.website, obj.website)
        return "-"
    website_link.short_description = "Website"
    
    def product_count(self, obj):
        """Count products for this brand"""
        return obj.products.count()
    product_count.short_description = "Products"
    
    def make_active(self, request, queryset):
        count = queryset.update(is_active=True)
        self.message_user(request, f"✅ {count} brand(s) activated successfully.")
    make_active.short_description = "Mark selected as active"
    
    def make_inactive(self, request, queryset):
        count = queryset.update(is_active=False)
        self.message_user(request, f"✅ {count} brand(s) deactivated successfully.")
    make_inactive.short_description = "Mark selected as inactive"
    
    def make_featured(self, request, queryset):
        count = queryset.update(is_featured=True)
        self.message_user(request, f"✅ {count} brand(s) marked as featured.")
    make_featured.short_description = "Mark selected as featured"
    
    def remove_featured(self, request, queryset):
        count = queryset.update(is_featured=False)
        self.message_user(request, f"✅ {count} brand(s) removed from featured.")
    remove_featured.short_description = "Remove selected from featured"
    


# -----------------------------
# ModelNumber Admin
# -----------------------------
@admin.register(ModelNumber)
class ModelNumberAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    ALL_LIST_COLUMNS = [
        ('code', 'Code'),
        ('name', 'Name'),
        ('brand', 'Brand'),
        ('is_active', 'Active'),
        ('product_count', 'Products'),
        ('created_at', 'Created'),
    ]
    DEFAULT_COLUMNS = ['code', 'name', 'brand', 'is_active', 'product_count', 'created_at']
    REQUIRED_COLUMNS = ['code']
    list_display = ('code', 'name', 'brand', 'is_active', 'product_count', 'created_at')
    list_filter = ('brand', 'is_active', 'created_at')
    search_fields = ('name', 'code', 'brand__name')
    autocomplete_fields = ['brand']
    list_editable = ('is_active',)
    readonly_fields = ('created_at', 'updated_at', 'product_count')  # product_count is a method, that's fine
    list_per_page = 25
    actions = ['make_active', 'make_inactive']
    export_fields = ['brand__name', 'name', 'code', 'is_active', 'created_at']
    export_methods = {
        'brand__name': lambda obj: obj.brand.name if obj.brand else '',
        'is_active': lambda obj: 'Yes' if obj.is_active else 'No',
    }

    fieldsets = (
        (None, {
            'fields': ('brand', 'name', 'code', 'description')
        }),
        ('Specifications', {
            'fields': ('specifications',),
            'classes': ('collapse',),
            'description': 'Store technical specifications in JSON format'
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
        ('Audit', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def product_count(self, obj):
        """Count products for this model number"""
        return obj.products.count()
    product_count.short_description = "Products"
    
    def make_active(self, request, queryset):
        count = queryset.update(is_active=True)
        self.message_user(request, f"✅ {count} model number(s) activated successfully.")
    make_active.short_description = "Mark selected as active"
    
    def make_inactive(self, request, queryset):
        count = queryset.update(is_active=False)
        self.message_user(request, f"✅ {count} model number(s) deactivated successfully.")
    make_inactive.short_description = "Mark selected as inactive"
    


# -----------------------------
# Product Category Admin (MPTT)
# -----------------------------


@admin.register(ProductCategory)
class ProductCategoryAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, DraggableMPTTAdmin):
    mptt_indent_field = "name"
    ALL_LIST_COLUMNS = [
        ('tree_actions', 'Actions'),
        ('indented_title', 'Category'),
        ('code', 'Code'),
        ('active', 'Active'),
        ('is_featured', 'Featured'),
        ('position', 'Position'),
        ('product_count', 'Products'),
        ('get_total_products', 'Total Products'),
    ]
    DEFAULT_COLUMNS = ['tree_actions', 'indented_title', 'code', 'active', 'is_featured', 'position', 'product_count']
    REQUIRED_COLUMNS = ['tree_actions', 'indented_title']
    list_display = ('tree_actions', 'indented_title', 'code', 'active',
                   'is_featured', 'position', 'product_count', 'get_total_products')
    list_display_links = ('indented_title',)
    search_fields = ('name', 'slug', 'code', 'parent__name')
    list_filter = ('active', 'is_featured', 'created_at')
    readonly_fields = ('created_at', 'updated_at', 'lft', 'rght', 'tree_id', 'level',
                       'product_count', 'get_total_products')  # Methods are fine here
    list_editable = ('position', 'active', 'is_featured')
    list_per_page = 25
    actions = ['make_active', 'make_inactive']
    export_fields = ['name', 'code', 'active', 'is_featured', 'default_discount', 'tax_rate', 'created_at']
    export_methods = {
        'active': lambda obj: 'Yes' if obj.active else 'No',
        'is_featured': lambda obj: 'Yes' if obj.is_featured else 'No',
    }

    fieldsets = (
        (None, {
            'fields': ('name', 'slug', 'parent', 'description', 'image', 'icon',
                      'color', 'position', 'active', 'is_featured', 'notes')
        }),
        ('Business / Code', {
            'fields': ('code', 'default_discount', 'tax_rate'),
        }),
        ('Statistics', {
            'fields': ('product_count', 'get_total_products'),
            'classes': ('collapse',)
        }),
        ('Tree Information', {
            'fields': ('lft', 'rght', 'tree_id', 'level'),
            'classes': ('collapse',)
        }),
        ('Audit', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('parent').annotate(
            product_count_annotated=Count('products')
        )
    
    def product_count(self, obj):
        """Count direct products in this category"""
        return obj.products.count()
    product_count.short_description = "Direct Products"
    product_count.admin_order_field = 'product_count_annotated'
    
    def get_total_products(self, obj):
        """Get total products including subcategories"""
        return obj.get_total_products()
    get_total_products.short_description = "Total Products (incl. subcategories)"
    
    def make_active(self, request, queryset):
        count = queryset.update(active=True)
        self.message_user(request, f"✅ {count} categories activated successfully.")
    make_active.short_description = "Mark selected as active"
    
    def make_inactive(self, request, queryset):
        count = queryset.update(active=False)
        self.message_user(request, f"✅ {count} categories deactivated successfully.")
    make_inactive.short_description = "Mark selected as inactive"
    


# -----------------------------
# Product Image Inline
# -----------------------------
class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    fields = ['image', 'alt_text', 'is_primary', 'position', 'image_preview']
    readonly_fields = ['image_preview']
    
    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="max-height: 50px;" />', obj.image.url)
        return "No image"
    image_preview.short_description = "Preview"


# -----------------------------
# Product Packing Inline
# -----------------------------
class ProductPackingInline(admin.TabularInline):
    model = ProductPacking
    extra = 1
    fields = ['packing_unit', 'quantity', 'price', 'is_default']
    autocomplete_fields = ['packing_unit']

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'packing_unit':
            obj_id = request.resolver_match.kwargs.get('object_id')
            if obj_id:
                used_ids = list(
                    ProductPacking.objects.filter(product_id=obj_id)
                    .values_list('packing_unit_id', flat=True)
                )
                qs = Unit.objects.filter(
                    Q(is_active=True) | Q(pk__in=used_ids)
                ) if used_ids else Unit.objects.filter(is_active=True)
            else:
                qs = Unit.objects.filter(is_active=True)
            kwargs['queryset'] = qs
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


# -----------------------------
# Product Variant Inline
# -----------------------------
class ProductVariantInline(admin.TabularInline):
    model = ProductVariant
    extra = 0
    fields = ['name', 'sku', 'price_adjustment', 'stock_quantity', 'is_active']
    readonly_fields = ['sku']


# -----------------------------
# Product Attribute Assignment Inline
# -----------------------------
class ProductAttributeAssignmentInline(admin.TabularInline):
    model = ProductAttributeAssignment
    extra = 1
    autocomplete_fields = ['attribute', 'value']


# -----------------------------
# Product Price History Inline
# -----------------------------
class ProductPriceHistoryInline(admin.TabularInline):
    model = ProductPriceHistory
    extra = 0
    readonly_fields = ['old_price', 'new_price', 'changed_by', 'changed_at', 'reason']
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False


# -----------------------------
# Product Admin - CLEANED VERSION (No inventory fields)
# -----------------------------
@admin.register(Product)
class ProductAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    ALL_LIST_COLUMNS = [
        ('name', 'Name'),
        ('brand', 'Brand'),
        ('category', 'Category'),
        ('sku', 'SKU'),
        ('price', 'Price'),
        ('discount_price', 'Discount Price'),
        ('product_type', 'Type'),
        ('active', 'Active'),
        ('is_featured', 'Featured'),
        ('image_preview', 'Image'),
    ]
    DEFAULT_COLUMNS = ['name', 'brand', 'category', 'sku', 'price', 'product_type', 'active', 'is_featured']
    REQUIRED_COLUMNS = ['name']
    list_display = (
        'name', 'brand', 'category', 'sku', 'price', 'discount_price',
        'product_type', 'active', 'is_featured', 'image_preview'
    )
    
    list_filter = (
        'brand',
        'category',
        'product_type',
        'active',
        'is_featured',
        'visibility',
        HasVariantsFilter,
        PriceRangeFilter,
        ('base_unit', admin.RelatedOnlyFieldListFilter),
        'created_at',
    )
    
    search_fields = ('name', 'sku', 'barcode', 'description', 'category__name', 
                    'brand__name', 'model_number__code', 'model_name')
    
    readonly_fields = ('created_at', 'updated_at', 'sku', 'slug', 'price_history_link')
    
    inlines = [
        ProductImageInline, 
        ProductPackingInline, 
        ProductVariantInline,
        ProductAttributeAssignmentInline,
        ProductPriceHistoryInline,
        ProductPriceInline,
    ]
    
    autocomplete_fields = ['category', 'related_products', 'base_unit', 'brand', 'default_price_list']
    
    radio_fields = {'product_type': admin.HORIZONTAL, 'visibility': admin.VERTICAL}
    
    list_editable = ('price', 'active', 'is_featured')
    
    list_per_page = 25
    
    date_hierarchy = 'created_at'
    
    actions = [
        'make_active', 
        'make_inactive', 
        'make_featured', 
        'remove_featured', 
        'export_pricelist_excel',
        'bulk_set_retail_prices', 
        'bulk_set_wholesale_prices', 
        'bulk_set_distributor_prices', 
        'copy_prices_from_base'
    ]
    
    export_fields = [
        'name', 'sku', 'barcode', 'brand__name', 'model_number__code',
        'category__name', 'base_unit__name', 'price', 'discount_price', 
        'cost', 'product_type', 'active', 'is_featured', 'weight', 'dimensions',
        'created_at'
    ]
    
    export_methods = {
        'brand__name': lambda obj: obj.brand.name if obj.brand else '',
        'model_number__code': lambda obj: obj.model_number.code if obj.model_number else '',
        'category__name': lambda obj: obj.category.name if obj.category else '',
        'base_unit__name': lambda obj: obj.base_unit.name if obj.base_unit else '',
    }
    
    fieldsets = (
        ('Basic Info', {
            'fields': ('name', 'slug', 'category', 'related_products', 'description')
        }),
        ('Brand & Model', {
            'fields': ('brand', 'model_number', 'model_name'),
        }),
        ('Pricing', {
            'fields': ('currency', 'base_unit', 'price', 'discount_price', 'cost', 'base_price', 'default_price_list'),
        }),
        ('Type & Status', {
            'fields': ('product_type', 'active', 'is_featured', 'position', 'visibility')
        }),
        ('Media', {
            'fields': ('main_image',),
            'classes': ('collapse',)
        }),
        ('Identification', {
            'fields': ('sku', 'barcode'),
        }),
        ('Shipping & Packaging', {
            'fields': ('weight', 'dimensions', 'multi_pack'),
            'classes': ('collapse',)
        }),
        ('Price History', {
            'fields': ('price_history_link',),
            'classes': ('collapse',)
        }),
        ('Notes / Audit', {
            'fields': ('notes', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Filter model_number based on selected brand; restrict unit choices to active."""
        if db_field.name == 'base_unit':
            obj_id = request.resolver_match.kwargs.get('object_id')
            current_id = None
            if obj_id:
                try:
                    current_id = Product.objects.values_list(
                        'base_unit_id', flat=True
                    ).get(pk=obj_id)
                except Product.DoesNotExist:
                    pass
            if current_id:
                kwargs['queryset'] = Unit.objects.filter(
                    Q(is_active=True) | Q(pk=current_id)
                )
            else:
                kwargs['queryset'] = Unit.objects.filter(is_active=True)
            field = super().formfield_for_foreignkey(db_field, request, **kwargs)
            if field:
                field.label_from_instance = lambda u: (
                    f"{u} [Inactive — must be replaced]" if not u.is_active else str(u)
                )
            return field

        if db_field.name == "model_number":
            current_obj = None
            if hasattr(request, 'resolver_match') and request.resolver_match:
                if 'object_id' in request.resolver_match.kwargs:
                    try:
                        obj_id = request.resolver_match.kwargs['object_id']
                        current_obj = self.get_object(request, obj_id)
                    except:
                        pass
            
            brand_id = None
            if request.method == 'POST':
                brand_id = request.POST.get('brand')
            elif current_obj and current_obj.brand:
                brand_id = current_obj.brand_id
            
            qs = ModelNumber.objects.filter(is_active=True)
            
            if brand_id:
                qs = qs.filter(brand_id=brand_id)
            
            if current_obj and current_obj.model_number:
                qs = qs | ModelNumber.objects.filter(id=current_obj.model_number_id)
            
            kwargs['queryset'] = qs.distinct().order_by('name')
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)

        class ProductFormWithUnitValidation(form):
            def clean_base_unit(self_form):
                unit = self_form.cleaned_data.get('base_unit')
                if unit and not unit.is_active:
                    raise ValidationError(
                        f'"{unit}" is inactive. Please select an active unit.'
                    )
                return unit

        return ProductFormWithUnitValidation

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'category', 'base_unit', 'brand', 'model_number', 'default_price_list'
        ).prefetch_related(
            'gallery_images', 'packings', 'variants', 'price_list_prices__price_list'
        )
    
    def image_preview(self, obj):
        image = obj.get_primary_image()
        if image:
            return format_html('<img src="{}" style="max-height: 40px; max-width: 40px;" />', image.url)
        return "No image"
    image_preview.short_description = "Image"
    
    def price_history_link(self, obj):
        """Generate link to price history for this product"""
        try:
            count = obj.price_history.count()
            
            if count == 0:
                return "No price history"
            
            from django.urls import reverse
            url = reverse('admin:products_productpricehistory_changelist') + f'?product__id__exact={obj.id}'
            
            return format_html(
                '<a href="{}" style="font-weight: bold;">📊 View {} price change{}</a>',
                url,
                count,
                's' if count != 1 else ''
            )
        except Exception:
            return "Price history unavailable"
    price_history_link.short_description = "Price History"
    
    # Bulk Actions
    def make_active(self, request, queryset):
        count = queryset.update(active=True)
        self.message_user(request, f"✅ {count} product(s) activated.")
    make_active.short_description = "Mark selected as active"

    def make_inactive(self, request, queryset):
        count = queryset.update(active=False)
        self.message_user(request, f"✅ {count} product(s) deactivated.")
    make_inactive.short_description = "Mark selected as inactive"

    def make_featured(self, request, queryset):
        count = queryset.update(is_featured=True)
        self.message_user(request, f"✅ {count} product(s) marked as featured.")
    make_featured.short_description = "Mark selected as featured"

    def remove_featured(self, request, queryset):
        count = queryset.update(is_featured=False)
        self.message_user(request, f"✅ {count} product(s) removed from featured.")
    remove_featured.short_description = "Remove selected from featured"

    def bulk_set_retail_prices(self, request, queryset):
        retail_list = PriceList.objects.filter(applicable_to_retail=True).first()
        if not retail_list:
            self.message_user(request, "❌ No retail price list found", level='ERROR')
            return
        
        count = 0
        for product in queryset:
            ProductPrice.objects.update_or_create(
                product=product,
                price_list=retail_list,
                defaults={'price': product.base_price}
            )
            count += 1
        
        self.message_user(request, f"✅ Set retail prices for {count} products.")
    bulk_set_retail_prices.short_description = "Set retail prices"

    def bulk_set_wholesale_prices(self, request, queryset):
        wholesale_list = PriceList.objects.filter(applicable_to_wholesale=True).first()
        if not wholesale_list:
            self.message_user(request, "❌ No wholesale price list found", level='ERROR')
            return
        
        count = 0
        for product in queryset:
            discount = wholesale_list.default_discount_percentage
            if discount:
                price_amount = product.base_price * (1 - discount/100)
            else:
                price_amount = product.base_price * Decimal('0.85')
            
            ProductPrice.objects.update_or_create(
                product=product,
                price_list=wholesale_list,
                defaults={'price': price_amount}
            )
            count += 1
        
        self.message_user(request, f"✅ Set wholesale prices for {count} products.")
    bulk_set_wholesale_prices.short_description = "Set wholesale prices"

    def bulk_set_distributor_prices(self, request, queryset):
        distributor_list = PriceList.objects.filter(applicable_to_distributor=True).first()
        if not distributor_list:
            self.message_user(request, "❌ No distributor price list found", level='ERROR')
            return
        
        count = 0
        for product in queryset:
            discount = distributor_list.default_discount_percentage
            if discount:
                price_amount = product.base_price * (1 - discount/100)
            else:
                price_amount = product.base_price * Decimal('0.75')
            
            ProductPrice.objects.update_or_create(
                product=product,
                price_list=distributor_list,
                defaults={'price': price_amount}
            )
            count += 1
        
        self.message_user(request, f"✅ Set distributor prices for {count} products.")
    bulk_set_distributor_prices.short_description = "Set distributor prices"

    def copy_prices_from_base(self, request, queryset):
        count = 0
        for product in queryset:
            for price_list in PriceList.objects.filter(is_active=True):
                if price_list.discount_method == 'percentage':
                    discount = price_list.default_discount_percentage
                    new_price = product.base_price * (1 - discount/100)
                    ProductPrice.objects.update_or_create(
                        product=product,
                        price_list=price_list,
                        defaults={'price': new_price}
                    )
                    count += 1
        
        self.message_user(request, f"✅ Updated {count} prices.")
    copy_prices_from_base.short_description = "Calculate prices from base"
    
    def export_pricelist_excel(self, request, queryset):
        """Special export for price list"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Price List"
            
            headers = ['SKU', 'Product Name', 'Brand', 'Model', 'Category', 'Base Unit', 
                      'Base Price', 'Retail Price', 'Wholesale Price', 'Distributor Price', 
                      'Cost', 'Status']
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            retail_list = PriceList.objects.filter(applicable_to_retail=True).first()
            wholesale_list = PriceList.objects.filter(applicable_to_wholesale=True).first()
            distributor_list = PriceList.objects.filter(applicable_to_distributor=True).first()
            
            for row, product in enumerate(queryset, 2):
                try:
                    ws.cell(row=row, column=1, value=product.sku or "")
                    ws.cell(row=row, column=2, value=product.name or "")
                    ws.cell(row=row, column=3, value=product.brand.name if product.brand else "")
                    ws.cell(row=row, column=4, value=product.model_number.code if product.model_number else "")
                    ws.cell(row=row, column=5, value=product.category.name if product.category else "")
                    ws.cell(row=row, column=6, value=product.base_unit.short_name if product.base_unit else "")
                    ws.cell(row=row, column=7, value=float(product.base_price) if product.base_price else 0)
                    
                    if retail_list:
                        retail_price = product.price_list_prices.filter(price_list=retail_list).first()
                        ws.cell(row=row, column=8, value=float(retail_price.price) if retail_price else "")
                    
                    if wholesale_list:
                        wholesale_price = product.price_list_prices.filter(price_list=wholesale_list).first()
                        ws.cell(row=row, column=9, value=float(wholesale_price.price) if wholesale_price else "")
                    
                    if distributor_list:
                        distributor_price = product.price_list_prices.filter(price_list=distributor_list).first()
                        ws.cell(row=row, column=10, value=float(distributor_price.price) if distributor_price else "")
                    
                    ws.cell(row=row, column=11, value=float(product.cost) if product.cost else 0)
                    ws.cell(row=row, column=12, value='Active' if product.active else 'Inactive')
                    
                except Exception as e:
                    print(f"Error exporting product {product.id}: {e}")
                    continue
            
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Price_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            wb.save(response)
            return response
            
        except Exception as e:
            self.message_user(request, f"❌ Price list export failed: {str(e)}", level='ERROR')
            return redirect(request.path)
    export_pricelist_excel.short_description = "Export as Multi-Tier Price List"

    class Media:
        js = ('products/js/dependent_dropdowns.js',)


# -----------------------------
# Product Attribute Admin
# -----------------------------
@admin.register(ProductAttribute)
class ProductAttributeAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    ALL_LIST_COLUMNS = [
        ('name', 'Name'),
        ('code', 'Code'),
        ('value_count', 'Values'),
    ]
    DEFAULT_COLUMNS = ['name', 'code', 'value_count']
    REQUIRED_COLUMNS = ['name']
    list_display = ['name', 'code', 'value_count']
    search_fields = ['name', 'code']
    inlines = []
    export_fields = ['name', 'code', 'description']
    
    def value_count(self, obj):
        return obj.values.count()
    value_count.short_description = "Values"


@admin.register(ProductAttributeValue)
class ProductAttributeValueAdmin(ERPAdminMixin, ConfigurableExportMixin, admin.ModelAdmin):
    list_display = ['attribute', 'value', 'code']
    list_filter = ['attribute']
    search_fields = ['value', 'code']
    autocomplete_fields = ['attribute']
    export_fields = ['attribute__name', 'value', 'code']


@admin.register(ProductVariant)
class ProductVariantAdmin(ERPAdminMixin, ColumnConfigMixin, ConfigurableExportMixin, admin.ModelAdmin):
    ALL_LIST_COLUMNS = [
        ('product', 'Product'),
        ('name', 'Variant Name'),
        ('sku', 'SKU'),
        ('effective_price', 'Price'),
        ('stock_quantity', 'Stock'),
        ('is_active', 'Active'),
    ]
    DEFAULT_COLUMNS = ['product', 'name', 'sku', 'effective_price', 'stock_quantity', 'is_active']
    REQUIRED_COLUMNS = ['product']
    list_display = ['product', 'name', 'sku', 'effective_price', 'stock_quantity', 'is_active']
    list_filter = ['is_active', 'product']
    search_fields = ['name', 'sku', 'product__name']
    readonly_fields = ['sku']
    list_per_page = 25
    export_fields = ['product__name', 'name', 'sku', 'price_adjustment',
                     'stock_quantity', 'is_active', 'attributes']


@admin.register(ProductPriceHistory)
class ProductPriceHistoryAdmin(ERPAdminMixin, ConfigurableExportMixin, admin.ModelAdmin):
    list_display = ('product', 'formatted_old_price', 'formatted_new_price', 
                   'price_difference', 'price_change_percent', 'changed_by', 
                   'changed_at', 'reason')
    list_filter = ('changed_at', 'changed_by')
    search_fields = ('product__name', 'product__sku', 'reason', 'product__barcode')
    readonly_fields = ('product', 'old_price', 'new_price', 'changed_by', 
                      'changed_at', 'reason', 'price_difference', 'price_change_percent')
    date_hierarchy = 'changed_at'
    list_per_page = 25

    fieldsets = (
        (None, {
            'fields': ('product', 'old_price', 'new_price', 'price_difference', 'price_change_percent')
        }),
        ('Details', {
            'fields': ('changed_by', 'changed_at', 'reason'),
        }),
    )
    
    export_fields = ['product__name', 'product__sku', 'old_price', 'new_price', 
                    'changed_by__username', 'changed_at', 'reason']
    
    def formatted_old_price(self, obj):
        return f"${obj.old_price}"
    formatted_old_price.short_description = 'Old Price'
    formatted_old_price.admin_order_field = 'old_price'
    
    def formatted_new_price(self, obj):
        return f"${obj.new_price}"
    formatted_new_price.short_description = 'New Price'
    formatted_new_price.admin_order_field = 'new_price'
    
    def price_difference(self, obj):
        diff = obj.new_price - obj.old_price
        if diff > 0:
            color = 'green'
            symbol = '▲'
        elif diff < 0:
            color = 'red'
            symbol = '▼'
        else:
            color = 'gray'
            symbol = '•'
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} ${}</span>',
            color,
            symbol,
            abs(diff)
        )
    price_difference.short_description = 'Difference'
    
    def price_change_percent(self, obj):
        if obj.old_price and obj.old_price != 0:
            percent = ((obj.new_price - obj.old_price) / obj.old_price) * 100
            if percent > 0:
                color = 'green'
                symbol = '↑'
            elif percent < 0:
                color = 'red'
                symbol = '↓'
            else:
                color = 'gray'
                symbol = '→'
            
            percent_str = "{:.1f}".format(abs(percent))
            return format_html(
                '<span style="color: {}; font-weight: bold;">{} {}%</span>',
                color,
                symbol,
                percent_str
            )
        return '-'
    price_change_percent.short_description = 'Change %'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product', 'changed_by')
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False