"""
Excel (.xlsx) export backend using openpyxl.
Ports and extends the logic from products/admin.py ExcelExportMixin.
"""
import io
from datetime import datetime, date
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    file_extension = 'xlsx'

    HEADER_COLOR = '366092'  # Dark blue — matches existing ERP style
    FOOTER_COLOR = 'D9E1F2'  # Light blue for footer row

    def export(self, admin_instance, queryset, options: dict) -> HttpResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        sheet_name = getattr(admin_instance, 'export_sheet_name', None) or admin_instance.model.__name__
        ws.title = sheet_name[:31]  # Excel sheet name limit

        include_headers = options.get('include_headers', True)
        include_footer = options.get('include_footer', False)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        fields = self._get_fields(admin_instance)
        date_fmt = options.get('date_format', '%Y-%m-%d')

        start_row = 1

        # Headers
        if include_headers:
            headers = self.get_headers(admin_instance)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            start_row += 1

        # Data rows
        rows = self.get_rows(admin_instance, queryset, options)
        for row_idx, row_data in enumerate(rows, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Number formatting
                if isinstance(value, float) and not isinstance(value, bool):
                    cell.number_format = '#,##0.00'

        # Footer row (summary)
        if include_footer and rows:
            footer_row = row_idx + 1 if rows else start_row
            footer_fill = PatternFill(start_color=self.FOOTER_COLOR, end_color=self.FOOTER_COLOR, fill_type='solid')
            footer_font = Font(italic=True)
            from django.utils import timezone as tz_module
            import datetime as dt_module
            footer_label = f"Exported {len(rows)} records on {dt_module.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            cell = ws.cell(row=footer_row, column=1, value=footer_label)
            cell.font = footer_font
            cell.fill = footer_fill
            if len(fields) > 1:
                ws.merge_cells(
                    start_row=footer_row, start_column=1,
                    end_row=footer_row, end_column=len(fields)
                )

        # Auto-size columns
        for col_idx in range(1, len(fields) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_num in range(1, min(len(rows) + start_row, 102)):
                try:
                    cell_val = ws[f'{col_letter}{row_num}'].value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Build response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{admin_instance.model.__name__}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type=self.content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buf.getvalue())
        return response
