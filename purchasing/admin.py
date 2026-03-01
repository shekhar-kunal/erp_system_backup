from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import redirect, render, get_object_or_404
from django import forms
from django.forms import inlineformset_factory
from django.utils.html import format_html
from django.db.models import F, Sum, Q
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.urls import reverse
from products.models import Product
from exports.mixins import ConfigurableExportMixin
from rbac.admin_mixins import ERPAdminMixin

from .models import (
    Vendor, PurchaseOrder, PurchaseOrderLine, 
    PurchasingSettings, PurchaseReceipt, PurchaseReceiptLine,
    PurchaseOrderHistory
)
from .forms import VendorAdminForm, PurchaseOrderLineForm, PurchaseReceiptLineForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.safestring import mark_safe




@admin.register(PurchaseOrderLine)
class PurchaseOrderLineAdmin(ERPAdminMixin, admin.ModelAdmin):
    list_display = ['id', 'order', 'product', 'quantity', 'price', 'received_quantity']
    list_filter = ['order__status', 'order__vendor']
    search_fields = ['order__po_number', 'product__name', 'product__sku']
    autocomplete_fields = ['product', 'unit', 'warehouse', 'section']
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'order', 'product', 'unit', 'warehouse', 'section'
        )


class PurchaseOrderLineInline(admin.TabularInline):
    model = PurchaseOrderLine
    form = PurchaseOrderLineForm
    extra = 1
    fields = [
        'product', 'quantity', 'unit', 'price', 
        'discount_percent', 'tax_rate', 'warehouse', 'section',
         'received_quantity'
    ]
    readonly_fields = ['subtotal', 'received_quantity']
    classes = ['collapse']
    
    def get_readonly_fields(self, request, obj=None):
        if obj and obj.status in ['confirmed', 'partial', 'done', 'cancelled']:
            return [f.name for f in self.model._meta.fields if f.name != 'id']
        return self.readonly_fields

    def has_change_permission(self, request, obj=None):
        if obj and obj.status in ['confirmed', 'partial', 'done', 'cancelled', 'rejected']:
            return False
        return True

    def has_add_permission(self, request, obj=None):
        if obj and obj.status in ['confirmed', 'partial', 'done', 'cancelled', 'rejected']:
            return False
        return True

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status in ['confirmed', 'partial', 'done', 'cancelled', 'rejected']:
            return False
        return True


class PurchaseReceiptLineInline(admin.TabularInline):
    model = PurchaseReceiptLine
    form = PurchaseReceiptLineForm
    extra = 1
    fields = [
        'order_line', 'product', 'quantity_received', 'quantity_rejected',
        'quality_status', 'batch_number', 'expiry_date', 'section'
    ]
    autocomplete_fields = ['order_line', 'product', 'section']
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "order_line" and hasattr(request, '_receipt_obj') and request._receipt_obj:
            kwargs["queryset"] = PurchaseOrderLine.objects.filter(
                order=request._receipt_obj.purchase_order
            ).select_related('product', 'order')
        elif db_field.name == "product":
            # Limit product queryset to those in the order lines
            if hasattr(request, '_receipt_obj') and request._receipt_obj:
                product_ids = PurchaseOrderLine.objects.filter(
                    order=request._receipt_obj.purchase_order
                ).values_list('product_id', flat=True).distinct()
                kwargs["queryset"] = Product.objects.filter(id__in=product_ids)
            else:
                kwargs["queryset"] = Product.objects.all()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'order_line', 'product', 'section'
        )


class PurchaseOrderHistoryInline(admin.TabularInline):
    model = PurchaseOrderHistory
    extra = 0
    fields = ['changed_at', 'changed_by', 'action', 'field_name', 'old_value', 'new_value']
    readonly_fields = fields
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False


@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(ERPAdminMixin, ConfigurableExportMixin, admin.ModelAdmin):
    inlines = [PurchaseOrderLineInline, PurchaseOrderHistoryInline]
    
    list_display = [
        'po_number', 'vendor_link', 'order_date', 'expected_date',
        'total_amount_colored', 'status_colored', 'receipt_status_bar',
        'action_buttons'
    ]
    list_filter = [
        'status', 'warehouse', 'order_date', 'expected_date',
        'vendor', 'payment_terms', 'currency'
    ]
    search_fields = ['po_number', 'vendor__name', 'vendor__code', 'vendor_reference']
    date_hierarchy = 'order_date'
    list_per_page = 25
    
    # Define fieldsets without non-editable fields
    fieldsets = (
        ('Basic Information', {
            'fields': (
                'po_number', 'vendor', 'warehouse', 'expected_date',
                'vendor_reference'
            )
        }),
        ('Financial Summary', {
            'fields': (
                ('subtotal', 'tax_amount', 'shipping_cost'),
                ('discount_amount', 'total_amount'),
                ('currency', 'exchange_rate', 'payment_terms')
            ),
            'classes': ('wide',)
        }),
        ('Shipping Information', {
            'fields': ('shipping_address', 'shipping_method', 'tracking_number'),
            'classes': ('collapse',)
        }),
        ('Additional Information', {
            'fields': ('notes', 'terms_conditions'),
            'classes': ('collapse',)
        }),
        ('Approval & Status', {
            'fields': (
                'status', 'created_by', 'approved_by', 'approved_at',
                'cancelled_by', 'cancelled_at', 'cancellation_reason'
            ),
            'classes': ('collapse',)
        }),
    )
    
    # Define which fields are read-only
    readonly_fields = [
        'po_number', 'subtotal', 'total_amount', 'created_by',
        'approved_by', 'approved_at', 'cancelled_by', 'cancelled_at',
        'order_date'  # Add order_date to readonly_fields since it's auto_now_add
    ]
    
    # Define which fields are not editable at all (won't appear in forms)
    exclude = []  # Don't exclude order_date, just make it readonly
    
    autocomplete_fields = ['vendor', 'warehouse']
    
    class Media:
        css = {
            'all': ('css/admin/purchasing.css',)
        }
        js = ('js/admin/purchase_order.js', 'js/dependent_dropdowns.js')

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'vendor', 'warehouse', 'created_by', 'approved_by'
        ).prefetch_related('lines')

    def get_readonly_fields(self, request, obj=None):
        """Dynamically set readonly fields"""
        readonly_fields = list(self.readonly_fields)
        
        # If object exists and is not in draft, make more fields readonly
        if obj and obj.status not in ['draft']:
            # Add more fields to readonly for non-draft orders
            additional_readonly = ['vendor', 'warehouse', 'expected_date', 
                                  'payment_terms', 'currency', 'shipping_address']
            for field in additional_readonly:
                if field not in readonly_fields:
                    readonly_fields.append(field)
        
        return readonly_fields

    def vendor_link(self, obj):
        return format_html(
            '<a href="{}">{}</a>',
            reverse('admin:purchasing_vendor_change', args=[obj.vendor.id]),
            obj.vendor.name
        )
    vendor_link.short_description = 'Vendor'
    vendor_link.admin_order_field = 'vendor__name'

    def total_amount_colored(self, obj):
        """Display total amount with color based on status"""
        color = 'green' if obj.status == 'done' else 'blue'
        html = f'<span style="color: {color}; font-weight: bold;">{obj.total_amount_display}</span>'
        return mark_safe(html)
    total_amount_colored.short_description = 'Total'
    total_amount_colored.admin_order_field = 'total_amount'

    def status_colored(self, obj):
        """Display status with color"""
        colors = {
            'draft': '#6c757d',      # gray
            'pending_approval': '#ffc107',  # yellow
            'approved': '#17a2b8',    # teal
            'confirmed': '#007bff',   # blue
            'partial': '#fd7e14',     # orange
            'done': '#28a745',        # green
            'cancelled': '#dc3545',   # red
            'rejected': '#dc3545',    # red
        }
        color = colors.get(obj.status, '#000')
        html = f'<span style="color: {color}; font-weight: bold;">{obj.get_status_display()}</span>'
        return mark_safe(html)
    status_colored.short_description = 'Status'
    status_colored.admin_order_field = 'status'

    def receipt_status_bar(self, obj):
        """Display receipt progress as a colored bar"""
        percentage = obj.receipt_status
        if percentage >= 100:
            color = '#28a745'
        elif percentage >= 50:
            color = '#ffc107'
        else:
            color = '#dc3545'
        
        # Fix: Convert percentage to string with format before passing to format_html
        percentage_str = f"{percentage:.1f}"
        
        return format_html(
            '<div style="width:100px; height:20px; background:#e9ecef; border-radius:4px; overflow:hidden;">'
            '<div style="width:{}%; height:100%; background:{}; text-align:center; line-height:20px; '
            'color:white; font-size:10px;">{}%</div>'
            '</div>',
            percentage, color, percentage_str
        )
    receipt_status_bar.short_description = 'Receipt Progress'
    receipt_status_bar.admin_order_field = 'status'  # Optional: allow sorting by status



    def action_buttons(self, obj):
        """Display action buttons for purchase orders"""
        buttons = []
        request = getattr(self, 'request', None)
        
        if obj.status == 'draft':
            url = reverse('admin:purchaseorder-confirm', args=[obj.id])
            buttons.append(
                f'<a class="button" href="{url}" '
                f'style="background-color:#28a745; color:white; padding:3px 8px; '
                f'border-radius:3px; text-decoration:none; margin:2px;">✓ Confirm</a>'
            )

        if obj.status in ['confirmed', 'partial']:
            url = reverse('admin:purchaseorder-receive', args=[obj.id])
            buttons.append(
                f'<a class="button" href="{url}" '
                f'style="background-color:#007bff; color:white; padding:3px 8px; '
                f'border-radius:3px; text-decoration:none; margin:2px;">📦 Receive</a>'
            )

        if obj.status in ['draft', 'confirmed', 'partial', 'approved']:
            url = reverse('admin:purchaseorder-cancel', args=[obj.id])
            buttons.append(
                f'<a class="button" href="{url}" '
                f'style="background-color:#dc3545; color:white; padding:3px 8px; '
                f'border-radius:3px; text-decoration:none; margin:2px;">✗ Cancel</a>'
            )

        if obj.status == 'pending_approval' and request and request.user.has_perm('purchasing.can_approve_purchase_orders'):
            url = reverse('admin:purchaseorder-approve', args=[obj.id])
            buttons.append(
                f'<a class="button" href="{url}" '
                f'style="background-color:#17a2b8; color:white; padding:3px 8px; '
                f'border-radius:3px; text-decoration:none; margin:2px;">✓ Approve</a>'
            )
        
        if buttons:
            return format_html('{}', '&nbsp;'.join(buttons))
        return format_html('<span>{}</span>', '—')
    action_buttons.short_description = 'Actions'

    def receipt_status_bar(self, obj):
        """Display receipt progress as a colored bar"""
        percentage = obj.receipt_status
        if percentage >= 100:
            color = '#28a745'
        elif percentage >= 50:
            color = '#ffc107'
        else:
            color = '#dc3545'
        
        pct_label = f'{percentage:.1f}%'
        return format_html(
            '<div style="width:100px; height:20px; background:#e9ecef; border-radius:4px; overflow:hidden;">'
            '<div style="width:{}%; height:100%; background:{}; text-align:center; line-height:20px; '
            'color:white; font-size:10px;">{}</div>'
            '</div>',
            percentage, color, pct_label
        )
    receipt_status_bar.short_description = 'Receipt Progress'

    def vendor_link(self, obj):
        """Link to vendor detail page"""
        url = reverse('admin:purchasing_vendor_change', args=[obj.vendor.id])
        return format_html('<a href="{}">{}</a>', url, obj.vendor.name)
    vendor_link.short_description = 'Vendor'
    vendor_link.admin_order_field = 'vendor__name'

    def total_amount_colored(self, obj):
        """Display total amount with color based on status"""
        color = 'green' if obj.status == 'done' else 'blue'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.total_amount_display
        )
    total_amount_colored.short_description = 'Total'
    total_amount_colored.admin_order_field = 'total_amount'

    def status_colored(self, obj):
        """Display status with color"""
        colors = {
            'draft': '#6c757d',
            'pending_approval': '#ffc107',
            'approved': '#17a2b8',
            'confirmed': '#007bff',
            'partial': '#fd7e14',
            'done': '#28a745',
            'cancelled': '#dc3545',
            'rejected': '#dc3545',
        }
        color = colors.get(obj.status, '#000')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_colored.short_description = 'Status'
    status_colored.admin_order_field = 'status'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('confirm/<int:order_id>/',
                 self.admin_site.admin_view(self.process_confirm),
                 name='purchaseorder-confirm'),
            path('receive/<int:order_id>/',
                 self.admin_site.admin_view(self.process_receive),
                 name='purchaseorder-receive'),
            path('cancel/<int:order_id>/',
                 self.admin_site.admin_view(self.process_cancel),
                 name='purchaseorder-cancel'),
            path('approve/<int:order_id>/',
                 self.admin_site.admin_view(self.process_approve),
                 name='purchaseorder-approve'),
            path('print/<int:order_id>/',
                 self.admin_site.admin_view(self.process_print),
                 name='purchaseorder-print'),
        ]
        return custom_urls + urls

    def save_model(self, request, obj, form, change):
        if not change:  # New object
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def save_formset(self, request, form, formset, change):
        """Save formset and update order totals"""
        instances = formset.save()
        if formset.model == PurchaseOrderLine:
            # Update order totals after line changes
            if hasattr(form, 'instance') and form.instance.pk:
                form.instance.calculate_totals()
                form.instance.save(update_fields=['subtotal', 'total_amount'])

    @transaction.atomic
    def process_confirm(self, request, order_id):
        order = get_object_or_404(PurchaseOrder, pk=order_id)
        
        if order.status != 'draft':
            self.message_user(request, "Only draft orders can be confirmed.", messages.ERROR)
            return redirect(f'../../{order_id}/change/')
        
        try:
            order.confirm()
            # Create history entry
            PurchaseOrderHistory.objects.create(
                purchase_order=order,
                changed_by=request.user,
                action='confirm',
                field_name='status',
                old_value='draft',
                new_value='confirmed'
            )
            self.message_user(request, f"Purchase Order {order.po_number} confirmed successfully.", messages.SUCCESS)
        except Exception as e:
            self.message_user(request, f"Error: {str(e)}", messages.ERROR)
        
        return redirect(f'../../{order_id}/change/')

    @transaction.atomic
    def process_approve(self, request, order_id):
        order = get_object_or_404(PurchaseOrder, pk=order_id)
        
        if order.status != 'pending_approval':
            self.message_user(request, "Only pending approval orders can be approved.", messages.ERROR)
            return redirect(f'../../{order_id}/change/')
        
        try:
            order.approve(request.user)
            self.message_user(request, f"Purchase Order {order.po_number} approved successfully.", messages.SUCCESS)
        except Exception as e:
            self.message_user(request, f"Error: {str(e)}", messages.ERROR)
        
        return redirect(f'../../{order_id}/change/')

    @transaction.atomic
    def process_cancel(self, request, order_id):
        order = get_object_or_404(PurchaseOrder, pk=order_id)
        
        if order.status not in ['draft', 'confirmed', 'partial', 'approved', 'pending_approval']:
            self.message_user(request, "This order cannot be cancelled.", messages.ERROR)
            return redirect(f'../../{order_id}/change/')
        
        if request.method == 'POST':
            reason = request.POST.get('reason', '')
            try:
                order.cancel(user=request.user, reason=reason)
                self.message_user(request, f"Purchase Order {order.po_number} cancelled successfully.", messages.SUCCESS)
                return redirect(f'../../{order_id}/change/')
            except Exception as e:
                self.message_user(request, f"Error: {str(e)}", messages.ERROR)
                return redirect(f'../../{order_id}/change/')
        
        # Show cancellation form
        context = {
            **self.admin_site.each_context(request),
            'title': f'Cancel Purchase Order - {order.po_number}',
            'order': order,
            'opts': self.model._meta,
        }
        return render(request, 'admin/purchasing/cancel_order.html', context)

    @transaction.atomic
    def process_receive(self, request, order_id):
        order = get_object_or_404(PurchaseOrder, pk=order_id)
        
        if order.status not in ['confirmed', 'partial']:
            self.message_user(request, "Only confirmed or partially received orders can receive goods.", messages.ERROR)
            return redirect(f'../../{order_id}/change/')

        # Create a receipt if it doesn't exist
        receipt, created = PurchaseReceipt.objects.get_or_create(
            purchase_order=order,
            status='draft',
            defaults={
                'received_by': request.user,
                'warehouse': order.warehouse
            }
        )

        # Store receipt in request for inline form
        request._receipt_obj = receipt

        ReceiptLineFormSet = inlineformset_factory(
            PurchaseReceipt,
            PurchaseReceiptLine,
            form=PurchaseReceiptLineForm,
            fields=['order_line', 'product', 'quantity_received', 'quantity_rejected',
                'quality_status', 'batch_number', 'expiry_date', 'section'],
            extra=order.lines.filter(quantity__gt=F('received_quantity')).count(),
            can_delete=False
        )

        if request.method == 'POST':
            formset = ReceiptLineFormSet(request.POST, instance=receipt)
            
            # STEP 1: Validate the formset
            if formset.is_valid():
                try:
                    with transaction.atomic():
                        # STEP 2: Save receipt lines
                        instances = formset.save(commit=False)
                        for instance in instances:
                            instance.receipt = receipt
                            instance.save()
                        
                        # STEP 3: Process each receipt line with validation
                        errors = []
                        for form in formset:
                            if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                                line_data = form.cleaned_data
                                order_line = line_data['order_line']
                                qty = line_data['quantity_received']
                                
                                # STEP 4: Double-check quantity validation
                                if qty > order_line.remaining_quantity:
                                    errors.append(
                                        f"{order_line.product.name}: Cannot receive {qty}, "
                                        f"only {order_line.remaining_quantity} remaining"
                                    )
                                    continue
                                
                                batch_info = {
                                    'batch_number': line_data.get('batch_number', ''),
                                    'expiry_date': line_data.get('expiry_date')
                                }
                                
                                # STEP 5: Call receive_line which has its own validation
                                try:
                                    order.receive_line(
                                        order_line.id, 
                                        qty, 
                                        batch_info=batch_info,
                                        quality_status=line_data.get('quality_status', 'accepted')
                                    )
                                except ValidationError as e:
                                    errors.append(str(e))
                        
                        # STEP 6: If there were errors, rollback and show them
                        if errors:
                            transaction.set_rollback(True)
                            for error in errors:
                                self.message_user(request, error, messages.ERROR)
                        else:
                            # STEP 7: Update order status from receipts after all lines processed
                            order.update_status_from_receipts()
                            
                            receipt.status = 'completed'
                            receipt.save()
                            self.message_user(request, "Goods received successfully.", messages.SUCCESS)
                            return redirect(f'../../{order_id}/change/')
                            
                except ValidationError as e:
                    self.message_user(request, f"Validation Error: {str(e)}", messages.ERROR)
                except Exception as e:
                    self.message_user(request, f"Error: {str(e)}", messages.ERROR)
            else:
                # STEP 8: Show form errors
                for error in formset.errors:
                    for field, msg in error.items():
                        self.message_user(request, f"{field}: {msg}", messages.ERROR)
        else:
            # Initialize formset with pending lines
            initial = []
            for line in order.lines.filter(quantity__gt=F('received_quantity')):
                initial.append({
                    'order_line': line,
                    'product': line.product,
                    'quantity_received': line.remaining_quantity,
                })
            formset = ReceiptLineFormSet(instance=receipt, initial=initial)

        context = {
            **self.admin_site.each_context(request),
            'title': f'Receive Goods – {order.po_number}',
            'order': order,
            'receipt': receipt,
            'formset': formset,
            'opts': self.model._meta,
        }
        return render(request, 'admin/purchasing/receive_goods.html', context)

    def process_print(self, request, order_id):
        order = get_object_or_404(PurchaseOrder, pk=order_id)
        
        # Simple HTML print version instead of PDF to avoid dependencies
        context = {
            **self.admin_site.each_context(request),
            'title': f'Print Purchase Order - {order.po_number}',
            'order': order,
            'opts': self.model._meta,
        }
        return render(request, 'admin/purchasing/print_order.html', context)
    
    export_fields = [
        'po_number', 'vendor__name', 'order_date', 'expected_date', 'status',
        'subtotal', 'tax_amount', 'shipping_cost', 'discount_amount', 'total_amount',
        'currency', 'payment_terms', 'warehouse__name', 'vendor_reference',
        'created_by__username', 'created_at',
    ]
    export_methods = {
        'vendor__name': lambda obj: obj.vendor.name if obj.vendor else '',
        'status': lambda obj: obj.get_status_display(),
        'payment_terms': lambda obj: obj.get_payment_terms_display(),
        'warehouse__name': lambda obj: obj.warehouse.name if obj.warehouse else '',
        'created_by__username': lambda obj: obj.created_by.username if obj.created_by else '',
    }
    export_sheet_name = 'Purchase Orders'

    actions = ['export_to_excel_detailed']

    def export_to_excel(self, request, queryset):
        """
        Export selected purchase orders to Excel with formatting
        """
        # Create a new workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Orders"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define headers
        headers = [
            'PO Number', 'Vendor', 'Order Date', 'Expected Date', 'Status',
            'Subtotal', 'Tax', 'Shipping', 'Discount', 'Total', 'Currency',
            'Payment Terms', 'Warehouse', 'Vendor Reference', 'Created By',
            'Created At', 'Items Count', 'Received Status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row, po in enumerate(queryset, 2):
            # Calculate items count and received status
            items_count = po.lines.count()
            received_percentage = po.receipt_status
            
            row_data = [
                po.po_number,
                po.vendor.name,
                po.order_date.strftime('%Y-%m-%d') if po.order_date else '',
                po.expected_date.strftime('%Y-%m-%d') if po.expected_date else '',
                po.get_status_display(),
                float(po.subtotal),
                float(po.tax_amount),
                float(po.shipping_cost),
                float(po.discount_amount),
                float(po.total_amount),
                po.currency,
                po.get_payment_terms_display(),
                po.warehouse.name if po.warehouse else '',
                po.vendor_reference or '',
                po.created_by.get_full_name() or po.created_by.username if po.created_by else '',
                po.created_at.strftime('%Y-%m-%d %H:%M') if po.created_at else '',
                items_count,
                f"{received_percentage:.1f}% Received"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 5:  # Status column
                    # Color code status
                    if po.status == 'done':
                        cell.font = Font(color="28a745")
                    elif po.status == 'cancelled':
                        cell.font = Font(color="dc3545")
                    elif po.status in ['confirmed', 'partial']:
                        cell.font = Font(color="007bff")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"purchase_orders_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        
        self.message_user(request, f"Exported {queryset.count()} purchase orders to Excel.", messages.SUCCESS)
        return response

    export_to_excel.short_description = "Export selected to Excel"

    def export_to_excel_detailed(self, request, queryset):
        """
        Export purchase orders with line items to Excel
        """
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Summary headers
        summary_headers = [
            'PO Number', 'Vendor', 'Order Date', 'Status', 'Total',
            'Items', 'Received %'
        ]
        
        # Style summary sheet
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write summary data
        for row, po in enumerate(queryset, 2):
            ws_summary.cell(row=row, column=1, value=po.po_number)
            ws_summary.cell(row=row, column=2, value=po.vendor.name)
            ws_summary.cell(row=row, column=3, value=po.order_date.strftime('%Y-%m-%d') if po.order_date else '')
            ws_summary.cell(row=row, column=4, value=po.get_status_display())
            ws_summary.cell(row=row, column=5, value=float(po.total_amount))
            ws_summary.cell(row=row, column=6, value=po.lines.count())
            ws_summary.cell(row=row, column=7, value=f"{po.receipt_status:.1f}%")
        
        # Create individual sheets for each PO with line items
        for po in queryset:
            sheet_name = f"PO-{po.po_number[-15:]}"  # Truncate if too long
            ws = wb.create_sheet(title=sheet_name)
            
            # PO Header
            ws.merge_cells('A1:H1')
            header_cell = ws.cell(row=1, column=1, value=f"Purchase Order: {po.po_number}")
            header_cell.font = Font(bold=True, size=14)
            header_cell.alignment = Alignment(horizontal='center')
            
            # PO Information
            info_data = [
                ('Vendor:', po.vendor.name),
                ('Order Date:', po.order_date.strftime('%Y-%m-%d') if po.order_date else ''),
                ('Expected Date:', po.expected_date.strftime('%Y-%m-%d') if po.expected_date else ''),
                ('Status:', po.get_status_display()),
                ('Payment Terms:', po.get_payment_terms_display()),
                ('Currency:', po.currency),
            ]
            
            row_num = 3
            for label, value in info_data:
                ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row_num, column=2, value=value)
                row_num += 1
            
            # Line Items Header
            row_num += 1
            line_headers = ['Product', 'Quantity', 'Unit', 'Price', 'Discount %', 'Tax %', 'Net Price', 'Total']
            for col, header in enumerate(line_headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
            
            # Line Items Data
            for line in po.lines.all():
                row_num += 1
                ws.cell(row=row_num, column=1, value=line.product.name if line.product else '')
                ws.cell(row=row_num, column=2, value=float(line.quantity))
                ws.cell(row=row_num, column=3, value=line.unit.name if line.unit else '')
                ws.cell(row=row_num, column=4, value=float(line.price))
                ws.cell(row=row_num, column=5, value=float(line.discount_percent or 0))
                ws.cell(row=row_num, column=6, value=float(line.tax_rate or 0))
                ws.cell(row=row_num, column=7, value=float(line.net_price or 0))
                ws.cell(row=row_num, column=8, value=float(line.subtotal))
            
            # Totals
            row_num += 2
            ws.cell(row=row_num, column=7, value="Subtotal:").font = Font(bold=True)
            ws.cell(row=row_num, column=8, value=float(po.subtotal))
            
            row_num += 1
            ws.cell(row=row_num, column=7, value="Tax:").font = Font(bold=True)
            ws.cell(row=row_num, column=8, value=float(po.tax_amount))
            
            row_num += 1
            ws.cell(row=row_num, column=7, value="Shipping:").font = Font(bold=True)
            ws.cell(row=row_num, column=8, value=float(po.shipping_cost))
            
            row_num += 1
            ws.cell(row=row_num, column=7, value="Discount:").font = Font(bold=True)
            ws.cell(row=row_num, column=8, value=float(po.discount_amount))
            
            row_num += 1
            total_cell = ws.cell(row=row_num, column=7, value="TOTAL:")
            total_cell.font = Font(bold=True, size=12)
            ws.cell(row=row_num, column=8, value=float(po.total_amount)).font = Font(bold=True, size=12)
        
        # Remove the default sheet if we created additional ones
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"purchase_orders_detailed_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        self.message_user(request, f"Exported {queryset.count()} purchase orders with details to Excel.", messages.SUCCESS)
        return response

    export_to_excel_detailed.short_description = "Export selected with line items"


@admin.register(PurchaseReceipt)
class PurchaseReceiptAdmin(ERPAdminMixin, admin.ModelAdmin):
    inlines = [PurchaseReceiptLineInline]
    
    list_display = [
        'receipt_number', 'purchase_order_link', 'received_date',
        'warehouse', 'status', 'items_count'
    ]
    list_filter = ['status', 'warehouse', 'received_date']
    search_fields = ['receipt_number', 'purchase_order__po_number', 'delivery_note_number']
    readonly_fields = ['receipt_number', 'received_date']
    
    fieldsets = (
        ('Receipt Information', {
            'fields': (
                'receipt_number', 'purchase_order', 'received_by',
                'received_date', 'warehouse', 'status'
            )
        }),
        ('Delivery Details', {
            'fields': (
                'delivery_note_number', 'vehicle_number',
                'driver_name', 'driver_phone'
            ),
            'classes': ('collapse',)
        }),
        ('Additional Info', {
            'fields': ('notes',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'purchase_order', 'warehouse', 'received_by'
        )
    
    def purchase_order_link(self, obj):
        return format_html(
            '<a href="{}">{}</a>',
            reverse('admin:purchasing_purchaseorder_change', args=[obj.purchase_order.id]),
            obj.purchase_order.po_number
        )
    purchase_order_link.short_description = 'Purchase Order'
    
    def items_count(self, obj):
        return obj.lines.count()
    items_count.short_description = 'Items'
    
    def save_model(self, request, obj, form, change):
        if not obj.received_by_id:
            obj.received_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(Vendor)
class VendorAdmin(ERPAdminMixin, ConfigurableExportMixin, admin.ModelAdmin):
    form = VendorAdminForm
    export_fields = [
        'name', 'code', 'contact_person', 'email', 'phone', 'website',
        'country__name', 'city__name', 'tax_number', 'payment_terms', 'currency',
        'credit_limit', 'is_active', 'is_preferred', 'quality_rating', 'total_orders', 'created_at',
    ]
    export_methods = {
        'country__name': lambda obj: obj.country.name if obj.country else '',
        'city__name': lambda obj: obj.city.name if obj.city else '',
        'payment_terms': lambda obj: obj.get_payment_terms_display() if hasattr(obj, 'get_payment_terms_display') else obj.payment_terms,
        'is_active': lambda obj: 'Yes' if obj.is_active else 'No',
        'is_preferred': lambda obj: 'Yes' if obj.is_preferred else 'No',
    }

    list_display = [
        'name', 'code', 'email', 'phone', 'country', 'city',
        'is_preferred_badge', 'performance_badge', 'total_purchases_display', 'is_active'
    ]
    list_filter = ['country', 'is_active', 'is_preferred', 'payment_terms']
    search_fields = ['name', 'code', 'email', 'contact_person', 'tax_number']
    list_select_related = ['country', 'city']
    
    fieldsets = (
        ('Basic Information', {
            'fields': (
                'name', 'code', 'contact_person', 'email', 'phone', 'mobile', 'website'
            )
        }),
        ('Address', {
            'fields': (
                'address_line1', 'address_line2',
                'country', 'region', 'city', 'postal_code'
            )
        }),
        ('Tax & Registration', {
            'fields': ('tax_number', 'registration_number', 'gst_number'),
            'classes': ('collapse',)
        }),
        ('Financial', {
            'fields': (
                'payment_terms', 'credit_days', 'credit_limit', 
                'opening_balance', 'currency'
            ),
            'classes': ('collapse',)
        }),
        ('Performance Metrics', {
            'fields': ('average_delivery_days', 'quality_rating', 'total_orders'),
            'classes': ('collapse',),
            'description': 'Automatically updated based on purchase history'
        }),
        ('Status', {
            'fields': ('is_active', 'is_preferred', 'notes', 'created_by')
        }),
    )
    
    readonly_fields = [
        'average_delivery_days', 'quality_rating', 'total_orders',
        'created_by', 'created_at', 'updated_at'
    ]
    
    actions = ['mark_preferred', 'mark_active', 'mark_inactive']
    
    class Media:
        js = ('js/dependent_dropdowns.js', 'js/duplicate_confirmation.js')

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(
            total_purchases_sum=Sum('purchase_orders__total_amount', 
                                   filter=Q(purchase_orders__status='done'))
        )

    def is_preferred_badge(self, obj):
        """Display preferred status as a badge"""
        if obj.is_preferred:
            return format_html(
                '<span style="background-color:#28a745; color:white; padding:3px 8px; '
                'border-radius:3px;">{}</span>',
                '★ Preferred'
            )
        return format_html('<span>{}</span>', '—')
    is_preferred_badge.short_description = 'Preferred'

    def performance_badge(self, obj):
        """Display quality rating as stars"""
        if obj.quality_rating >= 4.5:
            return format_html('<span style="color:#28a745; font-weight:bold;">{}</span>', '★★★★★')
        elif obj.quality_rating >= 4.0:
            return format_html('<span style="color:#ffc107; font-weight:bold;">{}</span>', '★★★★☆')
        elif obj.quality_rating >= 3.0:
            return format_html('<span style="color:#fd7e14; font-weight:bold;">{}</span>', '★★★☆☆')
        elif obj.quality_rating > 0:
            return format_html('<span style="color:#dc3545; font-weight:bold;">{}</span>', '★★☆☆☆')
        return format_html('<span>{}</span>', '—')
    performance_badge.short_description = 'Rating'

    def total_purchases_display(self, obj):
        """Display total purchases with currency"""
        if hasattr(obj, 'total_purchases_sum') and obj.total_purchases_sum:
            amount = f'{obj.total_purchases_sum:.2f}'
            return format_html('{} {}', obj.currency, amount)
        return format_html('{} 0.00', obj.currency)
    total_purchases_display.short_description = 'Total Purchases'
    total_purchases_display.admin_order_field = 'total_purchases_sum'

    def save_model(self, request, obj, form, change):
        if not change:  # New object
            obj.created_by = request.user
        
        # Check for duplicate confirmation
        if 'confirm_duplicate' in request.POST and request.POST.get('confirm_duplicate') == 'on':
            obj._confirm_duplicate = True
        
        super().save_model(request, obj, form, change)

    def mark_preferred(self, request, queryset):
        updated = queryset.update(is_preferred=True)
        self.message_user(request, f'{updated} vendors marked as preferred.')
    mark_preferred.short_description = "Mark selected as preferred vendors"

    def mark_active(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} vendors activated.')
    mark_active.short_description = "Mark selected as active"

    def mark_inactive(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} vendors deactivated.')
    mark_inactive.short_description = "Mark selected as inactive"


@admin.register(PurchasingSettings)
class PurchasingSettingsAdmin(ERPAdminMixin, admin.ModelAdmin):
    list_display = ['id', 'updated_at', 'updated_by']
    
    fieldsets = (
        ('Feature Toggles', {
            'fields': (
                ('enable_multi_currency', 'enable_purchase_approval'),
                ('enable_batch_tracking', 'enable_quality_check'),
                ('enable_partial_receiving', 'enable_auto_po_number'),
                ('enable_vendor_credit_notes', 'enable_purchase_returns'),
                'enable_receipt_batching',
            )
        }),
        ('Approval Workflow', {
            'fields': ('approval_levels', 'require_approval_above'),
            'classes': ('collapse',)
        }),
        ('Default Settings', {
            'fields': ('default_payment_terms', 'default_warehouse'),
            'classes': ('collapse',)
        }),
        ('PO Numbering', {
            'fields': ('po_prefix', 'po_next_number', 'po_number_padding'),
            'classes': ('collapse',)
        }),
        ('Notifications', {
            'fields': ('notify_on_overdue', 'notify_days_before'),
            'classes': ('collapse',)
        }),
    )
    
    def has_add_permission(self, request):
        # Only allow one instance
        if PurchasingSettings.objects.exists():
            return False
        return True
    
    def save_model(self, request, obj, form, change):
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)


# Register models
admin.site.register(PurchaseOrderHistory)