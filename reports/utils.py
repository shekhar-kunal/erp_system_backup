"""
Report export utility.

Usage:
    from reports.utils import export_report
    return export_report('Sales Report', columns, rows, fmt='excel')
"""
import csv
import io
from decimal import Decimal

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.text import slugify


def _fmt_value(v):
    """Stringify a value for CSV/PDF output."""
    if v is None:
        return ''
    if isinstance(v, Decimal):
        return str(v.quantize(Decimal('0.01')))
    return str(v)


def export_report(title, columns, rows, fmt='excel', summary=None):
    """
    title   : str
    columns : list[str]
    rows    : list[list]  — values may be int/Decimal/str/None
    fmt     : 'excel' | 'csv' | 'pdf'
    summary : dict (optional, used in PDF footer)
    Returns : HttpResponse
    """
    filename = slugify(title) or 'report'

    if fmt == 'excel':
        return _export_excel(title, columns, rows, filename, summary)
    elif fmt == 'csv':
        return _export_csv(title, columns, rows, filename)
    elif fmt == 'pdf':
        return _export_pdf(title, columns, rows, filename, summary)
    else:
        return _export_excel(title, columns, rows, filename, summary)


def _export_excel(title, columns, rows, filename, summary):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _export_csv(title, columns, rows, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1a237e')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Header row
    header_fill = PatternFill('solid', fgColor='3949ab')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Data rows
    alt_fill = PatternFill('solid', fgColor='F3F4F6')
    for row_idx, row in enumerate(rows, start=3):
        is_alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if is_alt:
                cell.fill = alt_fill
            cell.border = border
            # Right-align numbers
            if isinstance(val, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal='right')
                if isinstance(val, Decimal) or (isinstance(val, float) and '.' in str(val)):
                    cell.number_format = '#,##0.00'

    # Summary rows
    if summary:
        ws.append([])
        for k, v in summary.items():
            row_vals = [k, v] + [''] * (len(columns) - 2)
            ws.append(row_vals)
            r = ws.max_row
            ws.cell(r, 1).font = Font(bold=True)

    # Auto-width columns
    for col_idx in range(1, len(columns) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def _export_csv(title, columns, rows, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([title])
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_fmt_value(v) for v in row])
    content = output.getvalue()

    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


def _export_pdf(title, columns, rows, filename, summary):
    str_rows = [[_fmt_value(v) for v in row] for row in rows]
    html = render_to_string('reports/pdf_report.html', {
        'title': title,
        'columns': columns,
        'rows': str_rows,
        'summary': summary or {},
    })
    try:
        import weasyprint
        pdf_bytes = weasyprint.HTML(string=html).write_pdf()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
    except ImportError:
        return _export_csv(title, columns, rows, filename)
